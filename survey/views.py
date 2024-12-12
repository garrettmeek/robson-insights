import csv
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.utils.dateparse import parse_date
from django.shortcuts import get_object_or_404

from django.http import HttpResponse, JsonResponse
from django.core.mail import EmailMessage
from django.views import View
from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.authtoken.views import APIView

import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from datetime import datetime, timedelta
import re
from .serializers import EntrySerializer, FilterSerializer
from .models import Entry, Filter
from .permissions import CanReadEntry
from users.models import Group, UserProfile

class EntryListView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = EntrySerializer

    def get_queryset(self):
        user_profiles = UserProfile.objects.filter(
            user=self.request.user
        ).filter(Q(can_view=True) | Q(is_admin=True))
        allowed_groups = user_profiles.values_list('group', flat=True)
        return Entry.objects.filter(groups__in=allowed_groups).distinct()

    def perform_create(self, serializer):
        # Fetch all user profiles where the user has can_add permissions
        user_profiles = UserProfile.objects
        if not user_profiles.exists():
            raise PermissionDenied("You do not have permission to add entries to any group.")

        # Collect all groups the user can add to
        allowed_groups = user_profiles.values_list('group', flat=True)

        # Save the entry and associate it with all allowed groups
        entry = serializer.save(user=self.request.user)
        entry.groups.set(allowed_groups)


    def post(self, request, *args, **kwargs):
        if 'file' in request.FILES:
            return self.upload_file(request.FILES['file'])
        else:
            return super().post(request, *args, **kwargs)

    def upload_file(self, file):
        try:
            _, file_extension = os.path.splitext(file.name)
            df = pd.DataFrame()
            if file_extension == '.csv':
                with open(file, 'r') as file:
                    lines = file.readlines()

                start_row = next(i for i, line in enumerate(lines) if line.strip().lower().startswith("group"))
                df = pd.read_csv(file, skiprows=start_row, header=None)

            elif file_extension == '.xlsx':
                workbook = load_workbook(file, read_only=True)
                sheet = workbook.active

                start_row = next(i for i, row in enumerate(sheet.iter_rows(values_only=True))
                                if str(row[0]).strip().lower().startswith("group"))
                df = pd.read_excel(file, skiprows=start_row, header=None)
            i = 1
            count = 0

            try:
                if df.iat[0, i][:7] != "Quarter":
                    raise Exception
            except:
                return Response({'error': 'Invalid format'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

            while df.iat[0, i][:7] == "Quarter":
                date = min(datetime.strptime(re.sub(r'(\d+)(st|nd|rd|th)', r'\1', df.iat[0, i].split("- ")[1]), '%d %B %Y'), datetime.now())
                j = 2

                try:
                    if df.iat[j, 0][:5] != "Group":
                        raise Exception
                except:
                    return Response({'error': 'Invalid format'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

                while df.iat[j, 0][:5] == "Group":
                    v = df.iat[j, i]
                    if not pd.isna(v):
                        for k in range(df.iat[j, i]):
                            self._create_entry(df.iat[j, 0].split(' ')[1], 'n', date)
                            count += 1
                    c = df.iat[j, i + 1]
                    if not pd.isna(c):
                        for k in range(df.iat[j, i + 1]):
                            df.iat[j, 0].split(' ')[1], 'y', date
                            self._create_entry(df.iat[j, 0].split(' ')[1], 'y', date)
                            count += 1
                    j += 1
                i += 2
            return Response({"message": f"{count} entries uploaded successfully."}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _create_entry(self, classification, csection, date):
        try:
            user = self.request.user
            groups = UserProfile.objects.filter(user=user).values_list('group__name', flat=True).distinct()
            entry_data = {
                'classification': classification,
                'csection': csection,
                'date': date,
                'group': groups,
            }
            serializer = self.get_serializer(data=entry_data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class FilterEntriesByDateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        current_user = request.user

        user_groups = UserProfile.objects.filter(user=current_user).values_list('group', flat=True)

        if not user_groups.exists():
            return JsonResponse({'entries': []})

        start_date = request.POST.get('start_date', None)
        end_date = request.POST.get('end_date', None)

        try:
            start_date = parse_date(start_date) if start_date else None
            end_date = parse_date(end_date) if end_date else None
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)

        if start_date:
            start_date = datetime.combine(start_date, datetime.min.time())
        if end_date:
            end_date = datetime.combine(end_date, datetime.max.time())
        
        if start_date and end_date:
            entries = Entry.objects.filter(
                date__range=(start_date, end_date),
                groups__in=user_groups
            ).distinct()
        elif start_date:
            entries = Entry.objects.filter(
                date__gte=start_date,
                groups__in=user_groups
            ).distinct()
        elif end_date:
            entries = Entry.objects.filter(
                date__lte=end_date,
                groups__in=user_groups
            ).distinct()
        else:
            entries = Entry.objects.filter(groups__in=user_groups).distinct()
            
        entries = entries.order_by('date')

        entries_data = [
            {
                'id': entry.pk,
                'classification': entry.classification,
                'user': entry.user.username if entry.user else None,
                'groups': [group.name for group in entry.groups.all()],
                'csection': entry.csection,
                'date': entry.date.isoformat(),
            }
            for entry in entries
        ]

        return JsonResponse({'entries': entries_data})


class EntryFilterListView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = EntrySerializer

    def get_serializer(self, *args, **kwargs):
        # Exclude 'groups' field from the serializer
        kwargs['exclude_groups'] = True
        return super().get_serializer(*args, **kwargs)

    def get_queryset(self):
        pk = self.kwargs.get('pk')

        # Determine if the pk is prefixed with 'filter-' or 'group-'
        if pk.startswith('filter-'):
            filter_id = pk.split('-')[1]
            return self.get_entries_by_filter(filter_id)
        elif pk.startswith('group-'):
            group_id = pk.split('-')[1]
            return self.get_entries_by_group(group_id)
        else:
            return Entry.objects.none()

    def get_entries_by_filter(self, filter_id):
        try:
            user_filter = Filter.objects.get(pk=filter_id, user=self.request.user)
            user_profiles = UserProfile.objects.filter(
                user=self.request.user
            ).filter(Q(can_view=True) | Q(is_admin=True))
            allowed_groups = user_profiles.values_list('group', flat=True)

            groups_in_filter = user_filter.groups.filter(id__in=allowed_groups)
            return Entry.objects.filter(groups__in=groups_in_filter).distinct()

        except Filter.DoesNotExist:
            return Entry.objects.none()

    def get_entries_by_group(self, group_id):
        try:
            user_profiles = UserProfile.objects.filter(
                user=self.request.user
            ).filter(Q(can_view=True) | Q(is_admin=True), group__id=group_id)

            if not user_profiles.exists():
                raise PermissionDenied("You do not have permission to view entries for this group.")

            return Entry.objects.filter(groups__id=group_id).distinct()

        except Group.DoesNotExist:
            return Entry.objects.none()

class DownloadSurveyCSVView(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = EntrySerializer

    def get(self, request):
        try:
            queryset = Entry.objects.filter(
                user__in=UserProfile.objects.filter(
                    group__in=UserProfile.objects.filter(
                        user=request.user, can_view=True
                    ).values_list('group', flat=True)
                ).values_list('user', flat=True).distinct()
            )
            model = queryset.model
            model_fields = model._meta.fields + model._meta.many_to_many
            field_names = [field.name for field in model_fields]

            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="survey_data.csv"'
            writer = csv.writer(response, delimiter=",")
            writer.writerow(field_names)

            for row in queryset:
                values = [getattr(row, field) for field in field_names if field != "groups"]
                group_names = ', '.join(row.groups.values_list('name', flat=True))
                values.append(group_names)
                writer.writerow(values)

            recipient_email = request.GET.get('email')
            if recipient_email:
                email = EmailMessage(
                    subject='Survey Data CSV',
                    body='Please see the attached survey data.',
                    to=[recipient_email]
                )
                email.attach('survey_data.csv', response.getvalue(), 'text/csv')
                email.send()

                return Response({'message': 'CSV sent to email successfully!'}, status=status.HTTP_200_OK)

            return response

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EntryDetailView(generics.RetrieveAPIView):
    permission_classes = [CanReadEntry]
    serializer_class = EntrySerializer

    def get_queryset(self):

        pk = self.kwargs.get('pk')
        queryset = Entry.objects.get(pk=pk)
        return queryset


class FilterConfigurationListCreateView(generics.ListCreateAPIView):
    serializer_class = FilterSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user_filters = Filter.objects.filter(user=self.request.user)
        # Get groups the user has permission to view
        user_profiles = UserProfile.objects.filter(user=self.request.user, can_view=True)
        allowed_groups = user_profiles.values_list('group', flat=True)

        # Include filters with no groups or with allowed groups
        return user_filters.filter(Q(groups__in=allowed_groups) | Q(groups__isnull=True)).distinct()

    def perform_create(self, serializer):
        groups = serializer.validated_data.get('groups', [])

        user_groups = Group.objects.filter(userprofile__user=self.request.user)

        user_profiles = UserProfile.objects.filter(user=self.request.user, can_view=True)
        allowed_groups = user_profiles.values_list('group', flat=True)

        if not all(group in allowed_groups for group in groups):
            raise PermissionDenied("You can only add groups you belong to.")

        serializer.save(user=self.request.user)

class CreateConfiguration(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        configuration_name = request.data.get('configuration_name', '').strip()

        if not configuration_name:
            return Response({'error': 'Configuration name is required'}, status=status.HTTP_400_BAD_REQUEST)
        if len(configuration_name) > 100:
            return Response({'error': 'Group name cannot exceed 100 characters'}, status=status.HTTP_400_BAD_REQUEST)
        if len(configuration_name) < 5:
            return Response({'error': 'Group name must be at least 5 characters'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            if Filter.objects.filter(name__iexact=configuration_name).exists():
                return Response({'error': 'A group with this name already exists'}, status=status.HTTP_400_BAD_REQUEST)

            configuration = Filter.objects.create(name=configuration_name, user=self.request.user)

            return Response(
                {'success': f'Configuration "{configuration.name}" created.'},
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            print('Error:', str(e))
            return Response({'error': 'An unexpected error occurred.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class AddGroupToConfiguration(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        group_id = request.data.get('group_id')
        configuration_id = request.data.get('configuration_id')

        if not group_id or not configuration_id:
            return Response({'error': 'group_name and confuguration_id are required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            configuration = Filter.objects.get(id=configuration_id)
            group = Group.objects.get(id=group_id)
            configuration.groups.add(group)

            return Response({'success': f'{group.name} added to configuration {configuration.name}'}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class RemoveGroupFromConfiguration(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        group_id = request.data.get('group_id')
        configuration_id = request.data.get('configuration_id')

        if not group_id or not configuration_id:
            return Response({'error': 'group_id and confuguration_id are required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            configuration = Filter.objects.get(id=configuration_id)
            group = Group.objects.get(id=group_id)
            configuration.groups.remove(group)

            return Response({'success': f'{group.name} removed from configuration {configuration.name}'}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class FilterConfigurationDetailView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        pk = self.kwargs.get('pk')
        return Filter.objects.filter(pk=pk).prefetch_related('groups')

    def get(self, *args, **kwargs):
        pk = self.kwargs.get('pk')
        filter_instance = self.get_queryset().first()

        group_ids = list(filter_instance.groups.values_list('id', flat=True))

        return Response(group_ids)

class GenerateQuarterlyXLSX(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get_quarters(self):
        today = datetime.today()
        current_year_start = datetime(today.year, 7, 1)
        last_year_start = current_year_start - timedelta(days=365)

        if today >= current_year_start:
            start_date = last_year_start
        else:
            start_date = last_year_start - timedelta(days=365)

        quarters = [
            f"Quarter 1: 1st July {start_date.year} - 30th September {start_date.year}",
            f"Quarter 2: 1st October {start_date.year} - 31st December {start_date.year}",
            f"Quarter 3: 1st January {start_date.year + 1} - 31st March {start_date.year + 1}",
            f"Quarter 4: 1st April {start_date.year + 1} - 30th June {start_date.year + 1}",
        ]
        return quarters

    def get(self, request):
        quarters = self.get_quarters()

        group_labels = [
            "Group 1", "Group 2", "Group 3", "Group 4", "Group 5.1",
            "Group 5.2", "Group 6", "Group 7", "Group 8", "Group 9", "Group 10"
        ]

        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Quarterly Data"

        # Add headers
        headers = ["Group Robson"] + quarters + ["Final", ""]
        subheaders = [""] + ["Vaginal Delivery", "C/Section"] * (len(quarters) + 1)

        ws.append(headers)
        ws.append(subheaders)

        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(bold=True)

        # Merge each quarter header to span 2 columns
        for i in range(len(quarters)):
            # First merge quarter headers for each quarter (e.g., "Quarter 1: Vaginal Delivery" and "Quarter 1: C/Section")
            ws.merge_cells(start_row=1, start_column=2 + i * 2, end_row=1, end_column=3 + i * 2)
            ws.cell(row=1, column=2 + i * 2).value = quarters[i]
            ws.cell(row=1, column=2 + i * 2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=2 + i * 2).font = Font(bold=True)

        # Merge the "Final" header to span 2 columns
        ws.merge_cells(start_row=1, start_column=len(quarters) * 2 + 2, end_row=1, end_column=len(quarters) * 2 + 3)
        ws.cell(row=1, column=len(quarters) * 2 + 2).value = "Final"
        ws.cell(row=1, column=len(quarters) * 2 + 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=len(quarters) * 2 + 2).font = Font(bold=True)

        # Style subheaders (Vaginal Delivery and C/Section)
        for col in range(2, len(headers) * 2 + 1):
            ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Process data rows
        row_index = 3
        for group in group_labels:
            row = [group]
            # Leave data cells blank (users will input data)
            row.extend([0] * (len(quarters) * 2))
            # Add formulas for "Final" columns
            vaginal_formula = f"=SUMPRODUCT(B{row_index}:I{row_index}, --ISEVEN(COLUMN(B{row_index}:I{row_index})))"
            csection_formula = f"=SUMPRODUCT(B{row_index}:I{row_index}, --ISODD(COLUMN(B{row_index}:I{row_index})))"
            row.extend([vaginal_formula, csection_formula])
            ws.append(row)
            row_index += 1

        # Add "No Record" row
        ws.append(["No Record"] + [0] * (len(quarters) * 2) + [f"=SUMPRODUCT(B14:I14, --ISEVEN(COLUMN(B14:I14)))", f"=SUMPRODUCT(B14:I14, --ISODD(COLUMN(B14:I14)))"])
        row_index += 1

        # Add "Total" row with dynamic column formulas
        total_row = ["Total"]
        for col in range(2, len(headers) + 5):
            col_letter = get_column_letter(col)
            total_row.append(f"=SUM({col_letter}3:{col_letter}{row_index - 1})")
        ws.append(total_row)

        for col in range(1, len(headers) + 5):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = Font(bold=True)

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Save the workbook to the response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="quarterly_survey_data.xlsx"'
        wb.save(response)
        return response
    
    
class DeleteFilterView(APIView):
    
    def delete(self, request, pk):
        
        filter = get_object_or_404(Filter, pk=pk)
        
        try:
            filter = Filter.objects.get(pk=pk, user=request.user)
        except:
            return Response({'error': 'filter not found'}, status=status.HTTP_400_BAD_REQUEST)
        
        filter.delete()
        
        return Response({'success': 'filter deleted'}, status=status.HTTP_200_OK) 
